import csv, sqlite3
from openpyxl import load_workbook
import pprint

genre = 8 #Literatura - need to change this to make it select genre
status = "disponibila"

#connect to database
con = sqlite3.connect("Biblioteca.db")
c = con.cursor()

def get_rows(ws):
    for row in ws.iter_rows():
        yield [cell.value for cell in row]

def excel_to_db(book):  
    if book[0] and book[1]:
        author = book[0] + ' ' + book[1]
    elif book[1] is None:
        author = book[0]
    elif book[0] is None:
        author = book[1]
    else:
        author = ''

    db_book = (book[4], book[2], author, book[3], book[5], book[6], genre, status)
            #book_id, title, author, publishing, price, year, genre, status
    return db_book        


#read excel
wb = load_workbook(filename = r'Excels\Registru biblioteca mare 25.02.2020.xlsx', read_only = True)
ws=wb.active
#data needs to be read as a list of tuples
ex_books = list(get_rows(ws))
for ibx,book in enumerate(ex_books):
    if book[7] == 1:
        c.execute("INSERT OR IGNORE INTO book VALUES(?,?,?,?,?,?,?,?)", excel_to_db(book))
        print(book[4])
con.commit()
